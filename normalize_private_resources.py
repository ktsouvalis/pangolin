#!/usr/bin/env python3
"""
Normalize existing Pangolin private (site) resources to a naming convention
derived from their 10.x.y.z destination, and make sure each such resource
grants access to exactly the one user named in it.

Scope: any HOST or CIDR resource whose destination falls under 10.0.0.0/8.
Anything else (a destination outside 10.x.y.z entirely) doesn't fit this
per-owner convention and is left untouched (reported as SKIPPED).

  - host (a single 10.x.y.z address): "<city>-<username>-<vlan>-<z>", same
    as create_private_resources.py -- vlan = x concatenated with y
    zero-padded to 2 digits, z is the unpadded last octet.
  - cidr (10.x.y.z/prefix): "<city>-<username>-<vlan>-<tail>", where vlan is
    built from however many octets the prefix actually fixes (x if >=/16,
    x+y if >=/24, dropped entirely below /16), and tail is "all" for the
    common /24 case (matching the convention already in use) or the literal
    prefix length otherwise (e.g. "/16" -> tail "16"), so an unusual mask
    is never mistaken for a different /24. A prefix narrower than /16 (too
    broad to name a single owner's slice of it) is flagged instead of
    guessed.

For each in-scope resource:
  - city: taken from the existing name's first "-"-separated segment. The
    API has no separate "city" field -- this is the only source of truth,
    so a name with no "-" at all (no separable city segment) can't be
    normalized and is flagged instead of guessed.
  - target user (the one who should end up with access, and whose sanitized
    email local part becomes the username segment):
      * exactly 1 user currently assigned -> that user, unconditionally.
      * 2+ users assigned -> resolved ONLY if exactly one assigned user's
        sanitized local part matches a "-"-separated segment of the
        existing name (i.e. the name already appears to identify a primary
        owner among several grantees). Otherwise AMBIGUOUS, skipped.
      * 0 users assigned -> this is the expected state for a resource that
        create_private_resources.py created for an email that didn't match
        any org user yet (reported there as OK_NO_USER). Resolved (no
        confirmation needed) in this order:
          - name: peel the known city (name's first segment) and the known
            vlan/tail (computed from the destination, same as always) off
            the name, and if what's left is exactly one org user's
            sanitized local part, that's not a guess -- it's the literal
            reverse of create_private_resources.py's own naming formula.
          - niceId: same idea but from the other end -- peel the known
            vlan/tail and the resource's own (already-known) ports segment
            off the *end* of niceId, and if what's left at the front is
            exactly one org user's sanitized local part, that's the literal
            reverse of create_private_resources.py's niceId formula. Only
            attempted when tcpPortRangeString is a plain numeric list (see
            the niceId bullet above).
          - as of 2026-07-31: if both of the above fail, fall back to
            find_unique_segment_match() -- exactly one org user's sanitized
            local part appears as *any* standalone name segment, not
            necessarily in the specific position the exact reconstructions
            above require. Catches legacy/hand-created names that don't fit
            the computed city+vlan/tail shape at all. Still refuses to
            guess when more than one (or zero) org users match.
        Whichever resolves first is used. Access is granted
        (action=grant_access), no rename needed by construction. Still
        UNRESOLVED/skipped when the leftover/segment match is zero or
        ambiguous, with a best-effort suggested email for a human to
        confirm and apply by hand -- never auto-granted on an ambiguous
        match.
  - rename: POST /site-resource/{id} {"name": ...} if the computed name
    differs from the current one.
  - niceId: same POST /site-resource/{id} call (merged with the name fix
    when both apply) sets niceId to "<username>-<vlan>-<tail>-<ports>" --
    the target user's sanitized local part, the same vlan/tail segments as
    the name, and the resource's own tcpPortRangeString ports numerically
    sorted, each prefixed with "p" and dash-joined (e.g. "22-3389" ->
    "p22-p3389") -- mirroring create_private_resources.py's own niceId
    scheme (minus the city segment, which niceId never carried). Skipped
    (niceId left untouched) when tcpPortRangeString isn't a plain
    comma-separated numeric list (e.g. a "*" wildcard or a range) -- there's
    no ports segment to compute in that case. As of 2026-07-31,
    create_private_resources.py no longer sets niceId at all on creation
    (left to Pangolin's random default) -- this script is now the only
    place a deterministic niceId is ever assigned, for every resource
    regardless of how it was created.
  - access: POST /site-resource/{id}/users {"userIds": [target]} if the
    current user set isn't already exactly {target}. As of 2026-07-31, any
    *other* currently-assigned user with a resolvable email/userId is no
    longer just dropped by this call -- they're split off into their own
    new resource first (create_split_resource(): same destination/ports,
    spanning every org site, named/niceId'd for just them, mirroring
    create_private_resources.py's one-resource-per-user model) so pruning
    the original never silently revokes someone's access. A user with no
    resolvable email (rare, seen on some legacy resources) has no target to
    split to and is still just dropped. If any split creation fails, the
    original resource is left untouched rather than pruning access with
    nowhere for it to have gone.
  - Roles/clients are read (for the report) but never modified: every
    resource surveyed in this org -- both hand-created legacy ones and
    ones already produced by create_private_resources.py -- uniformly
    carries only the org's baseline "Admin" role and zero clients, which
    looks like server-side default behavior rather than per-resource state
    either script manages. A resource with anything else attached is
    flagged as an anomaly in the report rather than silently touched.
  - tcpPortRangeString/udpPortRangeString/disableIcmp are read but never
    intentionally changed by this script -- however, every POST update
    call re-asserts the resource's own current values for these explicitly.
    Confirmed live Pangolin API behavior: omitting them on an update doesn't
    leave them alone, it resets udpPortRangeString from blocked back to
    "all" on the resource being edited. Do not remove this echo-back to
    "simplify" the payload.

Defaults to a dry run: everything above is computed and reported, but no
API calls that change state are made. Pass --apply to actually rename
resources / prune access. A mass --apply (no --resource-id filter) asks
for interactive confirmation first unless --yes is also given.

Config (config.yml, same 'pangolin' block as create_private_resources.py):
    pangolin:
      base_url: "https://pangolin.uop.gr/int-api"
      org_slug: "university-of-the-peloponnese"
      api_key: "..."

Usage:
    python3 normalize_private_resources.py
    python3 normalize_private_resources.py --apply
    python3 normalize_private_resources.py --apply --resource-id 130 --resource-id 135
    python3 normalize_private_resources.py --apply --yes
"""

import argparse
import datetime
import sys

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

session = requests.Session()


def load_config(path):
    with open(path, "r") as f:
        cfg = yaml.safe_load(f)
    pg = cfg.get("pangolin", {})
    for key in ("base_url", "org_slug", "api_key"):
        if not pg.get(key):
            sys.exit(f"ERROR: missing 'pangolin.{key}' in {path}")
    pg["base_url"] = pg["base_url"].rstrip("/")
    return pg


def api_headers(cfg):
    return {"Authorization": f"Bearer {cfg['api_key']}", "Content-Type": "application/json"}


def verify_org(cfg):
    resp = session.get(f"{cfg['base_url']}/v1/org/{cfg['org_slug']}", headers=api_headers(cfg))
    resp.raise_for_status()
    org = resp.json()["data"]
    print(f"Org OK: {org.get('name', cfg['org_slug'])} (orgId={cfg['org_slug']})")
    return org


def get_all_pages(cfg, path, data_key, page_size=1000):
    items, page = [], 1
    while True:
        resp = session.get(f"{cfg['base_url']}/v1{path}", headers=api_headers(cfg),
                            params={"pageSize": page_size, "page": page})
        resp.raise_for_status()
        body = resp.json()
        page_items = body["data"][data_key]
        items.extend(page_items)
        total = body["data"].get("pagination", {}).get("total", len(items))
        page += 1
        if len(items) >= total or not page_items:
            break
    return items


def get_site_resources(cfg):
    return get_all_pages(cfg, f"/org/{cfg['org_slug']}/site-resources", "siteResources")


def build_org_email_index(cfg):
    """sanitized local-part -> [(email, userId)] across every org user, used
    both for the strict 0-user auto-resolve and for best-effort suggestions
    on resources with no resolvable currently-assigned user."""
    users = get_all_pages(cfg, f"/org/{cfg['org_slug']}/users", "users")
    index = {}
    for u in users:
        email = (u.get("email") or u.get("user", {}).get("email") or "").lower()
        uid = u.get("id") or u.get("user", {}).get("id")
        if not email or not uid:
            continue
        index.setdefault(sanitize_username(email), []).append((email, uid))
    return index


def get_resource_users(cfg, resource_id):
    resp = session.get(f"{cfg['base_url']}/v1/site-resource/{resource_id}/users",
                        headers=api_headers(cfg))
    resp.raise_for_status()
    return resp.json()["data"]["users"]


def get_resource_roles(cfg, resource_id):
    resp = session.get(f"{cfg['base_url']}/v1/site-resource/{resource_id}/roles",
                        headers=api_headers(cfg))
    resp.raise_for_status()
    return resp.json()["data"]["roles"]


def get_resource_clients(cfg, resource_id):
    resp = session.get(f"{cfg['base_url']}/v1/site-resource/{resource_id}/clients",
                        headers=api_headers(cfg))
    resp.raise_for_status()
    return resp.json()["data"]["clients"]


def update_resource(cfg, resource_id, fields):
    resp = session.post(f"{cfg['base_url']}/v1/site-resource/{resource_id}",
                         headers=api_headers(cfg), json=fields)
    resp.raise_for_status()


def set_resource_users(cfg, resource_id, user_ids):
    resp = session.post(f"{cfg['base_url']}/v1/site-resource/{resource_id}/users",
                         headers=api_headers(cfg), json={"userIds": user_ids})
    resp.raise_for_status()


def get_sites(cfg):
    """Same call create_private_resources.py makes -- every resource this
    tooling creates (from an xlsx row, or split off a multi-user legacy
    resource here) spans every site in the org for HA."""
    sites = get_all_pages(cfg, f"/org/{cfg['org_slug']}/sites", "sites")
    if not sites:
        sys.exit("ERROR: no sites found in this org.")
    return sites


def create_split_resource(cfg, sites, res, city, vlan, tail, ports, email, user_id):
    """Create a new site-resource for a user being split off a multi-user
    resource (see process_resource) -- same destination/ports/alias as the
    resource they're being split from, but named and owned for just them.
    Unlike create_private_resources.py's own rows, the target user is
    already fully resolved here, so niceId is set immediately rather than
    left random."""
    username = sanitize_username(email)
    name = f"{city}-{username}-{vlan}-{tail}"
    payload = {
        "name": name,
        "mode": res["mode"],
        "destination": res["destination"],
        "tcpPortRangeString": res.get("tcpPortRangeString") or "",
        "udpPortRangeString": "",
        "disableIcmp": True,
        "roleIds": [],
        "clientIds": [],
        "userIds": [user_id],
        "siteIds": [s["siteId"] for s in sites],
    }
    if ports is not None:
        payload["niceId"] = compute_expected_nice_id(username, vlan, tail, ports)
    if res.get("alias"):
        payload["alias"] = res["alias"]

    resp = session.put(f"{cfg['base_url']}/v1/org/{cfg['org_slug']}/site-resource",
                        headers=api_headers(cfg), json=payload)
    resp.raise_for_status()
    return resp.json()["data"]


def sanitize_username(email):
    return email.split("@", 1)[0].strip().lower().replace(".", "")


def compute_vlan_z(destination):
    """10.x.y.z only -> (vlan, z), e.g. 10.23.2.50 -> ("2302", "50"). Anything
    else (not 4 numeric octets, or first octet isn't literally "10") -> (None, None)."""
    parts = destination.strip().split(".")
    if len(parts) != 4 or not all(p.isdigit() for p in parts) or parts[0] != "10":
        return None, None
    x, y, z = parts[1], parts[2], parts[3]
    return f"{int(x)}{int(y):02d}", str(int(z))


def compute_cidr_vlan_tail(destination):
    """10.x.y.z/prefix only -> (vlan, tail), vlan built from whichever octets
    the prefix actually fixes (dropped below /16), tail "all" for the common
    /24 case or the literal prefix length otherwise, e.g.:
        10.15.25.0/24 -> ("1525", "all")   (x + zero-padded y; z is wildcard)
        10.15.0.0/16  -> ("15", "16")      (only x is fixed)
        10.0.0.0/8    -> (None, None)      (too broad to name a single owner)
    """
    ip_part, sep, prefix_str = destination.strip().partition("/")
    if not sep or not prefix_str.isdigit():
        return None, None
    prefix = int(prefix_str)
    parts = ip_part.split(".")
    if len(parts) != 4 or not all(p.isdigit() for p in parts) or parts[0] != "10":
        return None, None
    x, y, z = parts[1], parts[2], parts[3]
    fixed_octets = max(0, min(3, prefix // 8 - 1))  # octets beyond the leading "10" fully fixed
    if fixed_octets == 0:
        return None, None
    vlan_tokens = [str(int(x))]
    if fixed_octets >= 2:
        vlan_tokens.append(f"{int(y):02d}")
    if fixed_octets >= 3:
        vlan_tokens.append(str(int(z)))
    tail = "all" if prefix == 24 else str(prefix)
    return "".join(vlan_tokens), tail


def compute_naming_segments(destination, mode):
    """Dispatch to the host or cidr vlan/tail computation for `mode`."""
    if mode == "host":
        return compute_vlan_z(destination)
    if mode == "cidr":
        return compute_cidr_vlan_tail(destination)
    return None, None


def compute_expected_ports(tcp_port_range_string):
    """Parse a resource's tcpPortRangeString into sorted numeric port tokens
    for niceId purposes, or None if it isn't a plain comma-separated numeric
    list (e.g. a "*" wildcard, a range like "8000-9000", or blank) -- niceId
    normalization is skipped in that case, since there's no discrete ports
    segment to compute.
    """
    raw = (tcp_port_range_string or "").strip()
    if not raw:
        return None
    tokens = [p.strip() for p in raw.split(",") if p.strip()]
    if not tokens or not all(p.isdigit() for p in tokens):
        return None
    return sorted(tokens, key=int)


def compute_expected_nice_id(username, vlan, tail, ports):
    return "-".join([username, vlan, tail] + [f"p{p}" for p in ports])


def resolve_unassigned_target_from_nice_id(nice_id, vlan, tail, ports, org_email_index):
    """Strict reverse of create_private_resources.py's niceId formula
    ("<username>-<vlan>-<tail>-<ports>"), for a resource with 0 users
    currently assigned. Mirrors resolve_unassigned_target but peels from the
    *end* of niceId (vlan/tail/ports are already independently known) since
    niceId carries no city segment for a prefix. Returns (email, userId) or
    (None, None) if niceId isn't an exact match, ports couldn't be computed,
    or the leftover doesn't uniquely resolve.
    """
    if not nice_id or ports is None:
        return None, None
    suffix = "-" + "-".join([vlan, tail] + [f"p{p}" for p in ports])
    lnice = nice_id.lower()
    if not lnice.endswith(suffix):
        return None, None
    candidate_username = lnice[: -len(suffix)]
    if not candidate_username or "-" in candidate_username:
        return None, None
    candidates = org_email_index.get(candidate_username, [])
    if len(candidates) != 1:
        return None, None
    return candidates[0]


def resolve_unassigned_target(name, city, vlan, tail, org_email_index):
    """Strict reverse of create_private_resources.py's naming formula, for a
    resource with 0 users currently assigned. Peels the known city and known
    vlan/tail (both already independently derived -- city from the name's
    own first segment, vlan/tail from the destination) off the name; if
    what's left is exactly one org user's sanitized local part, that's not a
    guess, so it's safe to auto-grant. Returns (email, userId) or
    (None, None) if the name isn't an exact convention match or the leftover
    doesn't uniquely resolve.
    """
    if not name:
        return None, None
    lname = name.lower()
    prefix, suffix = f"{city}-", f"-{vlan}-{tail}"
    if not (lname.startswith(prefix) and lname.endswith(suffix)):
        return None, None
    candidate_username = lname[len(prefix):-len(suffix)]
    if not candidate_username or "-" in candidate_username:
        return None, None
    candidates = org_email_index.get(candidate_username, [])
    if len(candidates) != 1:
        return None, None
    return candidates[0]


def find_unique_segment_match(name, org_email_index):
    """Loosest resolution: exactly one org user's sanitized local part
    appears as a standalone "-"-separated segment of the name (not
    necessarily in the specific city/vlan/tail positions the strict
    reconstruction requires). Returns (email, userId), or (None, None) if
    zero or multiple org users match. Used both as the auto-apply fallback
    for 0-user resources (once the strict reconstruction fails) and as the
    best-effort suggestion surfaced for ambiguous 2+ user resources (never
    auto-applied there -- see resolve_target)."""
    if not name:
        return None, None
    name_segments = set(name.lower().split("-"))
    candidates = sorted({(e, uid) for seg in name_segments
                          for e, uid in org_email_index.get(seg, [])})
    return candidates[0] if len(candidates) == 1 else (None, None)


def resolve_target(name, users, org_email_index):
    """Return (target_email, target_user_id, reason_if_unresolved, suggested_email).

    users: current site-resource users (list of {userId, email, ...}).
    Only the first three are meaningful when resolution succeeds (reason is
    None); the last is only populated when resolution fails with 0 users.
    """
    name_segments = set(name.lower().split("-")) if name else set()
    resolved = [(u.get("email"), u.get("userId")) for u in users if u.get("email")]

    if len(users) == 0:
        suggested_email, _ = find_unique_segment_match(name, org_email_index)
        return None, None, "no users currently assigned", suggested_email

    if len(resolved) == 1 and len(users) == 1:
        email, uid = resolved[0]
        return email.lower(), uid, None, None

    matches = [(e, uid) for e, uid in resolved if sanitize_username(e) in name_segments]
    if len(matches) == 1:
        email, uid = matches[0]
        return email.lower(), uid, None, None

    emails_desc = ", ".join(u.get("email") or f"<no-email:{u.get('username', u.get('userId'))}>"
                             for u in users)
    return None, None, f"{len(users)} users assigned, ambiguous ({emails_desc})", None


def process_resource(cfg, sites, res, org_email_index, apply_changes):
    resource_id = res["siteResourceId"]
    row = {
        "resource_id": resource_id, "nice_id": res.get("niceId"), "new_nice_id": res.get("niceId"),
        "mode": res["mode"],
        "destination": res["destination"], "city": None, "target_email": None,
        "old_name": res["name"], "new_name": res["name"], "action": "none",
        "removed_users": None, "current_users": None, "roles": None, "clients": None,
        "status": None, "reason": None,
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    if res["mode"] not in ("host", "cidr"):
        row["status"] = "SKIPPED"
        row["reason"] = f"mode={res['mode']!r} -- only host/cidr resources are in scope"
        return row

    vlan, tail = compute_naming_segments(res["destination"], res["mode"])
    if vlan is None:
        row["status"] = "SKIPPED"
        row["reason"] = (f"destination {res['destination']!r} is not a 10.x.y.z {res['mode']} "
                          f"range (or the mask is too broad to name a single owner's slice)")
        return row

    name_parts = res["name"].split("-", 1) if res["name"] else []
    if len(name_parts) < 2 or not name_parts[0].strip():
        row["status"] = "SKIPPED"
        row["reason"] = f"name {res['name']!r} has no separable city segment"
        return row
    city = name_parts[0].strip().lower()
    row["city"] = city

    ports = compute_expected_ports(res.get("tcpPortRangeString"))

    users = get_resource_users(cfg, resource_id)
    row["current_users"] = ", ".join(u.get("email") or f"<no-email:{u.get('username')}>"
                                      for u in users) or "(none)"

    roles = get_resource_roles(cfg, resource_id)
    clients = get_resource_clients(cfg, resource_id)
    role_names = sorted(r["name"] for r in roles)
    row["roles"] = ", ".join(role_names) or "(none)"
    row["clients"] = str(len(clients))
    notes = []
    if role_names != ["Admin"] or clients:
        notes.append(f"unexpected roles/clients (roles={role_names}, clients={len(clients)}) "
                      f"-- not modified, review manually")
    if ports is None:
        notes.append(f"niceId not checked: tcpPortRangeString "
                      f"{res.get('tcpPortRangeString')!r} isn't a plain numeric list")

    if len(users) == 0:
        strict_email, strict_user_id = resolve_unassigned_target(
            res["name"], city, vlan, tail, org_email_index)
        if not strict_email:
            strict_email, strict_user_id = resolve_unassigned_target_from_nice_id(
                res.get("niceId"), vlan, tail, ports, org_email_index)
        if not strict_email:
            # As of 2026-07-31: the strict reconstruction only works when the
            # name/niceId already fits this convention's exact shape, which
            # legacy hand-created resources with no city/vlan/tail structure
            # never will. Falling back to "exactly one org user's sanitized
            # local part appears as a standalone name segment" (previously
            # only ever surfaced as a manual-review suggestion, see
            # resolve_target) catches those too, while still refusing to
            # guess when that match isn't unique.
            strict_email, strict_user_id = find_unique_segment_match(res["name"], org_email_index)
    else:
        strict_email, strict_user_id = None, None

    if strict_email:
        target_email, target_user_id, unresolved_reason, suggestion = (
            strict_email, strict_user_id, None, None)
    else:
        target_email, target_user_id, unresolved_reason, suggestion = resolve_target(
            res["name"], users, org_email_index)

    if unresolved_reason:
        row["status"] = "SKIPPED"
        row["reason"] = unresolved_reason + (f" -- org directory suggests: {suggestion}"
                                              if suggestion else "")
        if notes:
            row["reason"] += "; " + "; ".join(notes)
        return row

    row["target_email"] = target_email
    username = sanitize_username(target_email)
    expected_name = f"{city}-{username}-{vlan}-{tail}"
    row["new_name"] = expected_name

    expected_nice_id = compute_expected_nice_id(username, vlan, tail, ports) if ports is not None else None
    row["new_nice_id"] = expected_nice_id if expected_nice_id is not None else res.get("niceId")

    needs_rename = expected_name != res["name"]
    needs_nice_id_fix = expected_nice_id is not None and expected_nice_id != (res.get("niceId") or "")
    current_ids = {u.get("userId") for u in users if u.get("userId")}
    needs_access_fix = current_ids != {target_user_id}

    # As of 2026-07-31: a user other than the resolved target no longer just
    # loses access when this resource gets pruned down to one owner -- they
    # get their own new resource instead (same destination/ports, spanning
    # every org site, mirroring create_private_resources.py's one-resource-
    # per-user model), so multi-user legacy resources can be split apart
    # without silently dropping anyone's access. A user with no resolvable
    # email/userId (seen on some legacy resources) can't be split off (there's
    # no target to create a resource for) and is still just pruned.
    split_off_users = [(u.get("email").lower(), u.get("userId")) for u in users
                        if u.get("userId") != target_user_id and u.get("email") and u.get("userId")]
    unsplittable_users = [u.get("username") or u.get("userId") for u in users
                           if u.get("userId") != target_user_id and not u.get("email")]
    removed_desc = [f"{e} -> new resource" for e, _ in split_off_users]
    removed_desc += [f"<no-email:{u}> -> dropped (no email to split to)" for u in unsplittable_users]
    if removed_desc:
        row["removed_users"] = ", ".join(removed_desc)

    reason = "; ".join(notes) or None

    if not needs_rename and not needs_nice_id_fix and not needs_access_fix:
        row["status"] = "OK"
        row["reason"] = reason
        return row

    actions = []
    if needs_rename:
        actions.append("rename")
    if needs_nice_id_fix:
        actions.append("fix_nice_id")
    if needs_access_fix:
        actions.append("grant_access" if len(users) == 0 else
                        ("split_access" if split_off_users else "prune_access"))
    row["action"] = "+".join(actions)

    if not apply_changes:
        row["status"] = "DRY-RUN"
        row["reason"] = reason
        return row

    try:
        # Split off every other user into their own resource *before*
        # touching the original -- if any of these fails, the original is
        # left untouched rather than pruning someone's access with nowhere
        # for it to have gone.
        created = []
        for email, user_id in split_off_users:
            new_res = create_split_resource(cfg, sites, res, city, vlan, tail, ports, email, user_id)
            created.append(f"{email} -> siteResourceId={new_res['siteResourceId']} "
                            f"niceId={new_res['niceId']}")
        if created:
            row["removed_users"] = ", ".join(created + [
                f"<no-email:{u}> -> dropped (no email to split to)" for u in unsplittable_users
            ])

        update_fields = {}
        if needs_rename:
            update_fields["name"] = expected_name
        if needs_nice_id_fix:
            update_fields["niceId"] = expected_nice_id
        if update_fields:
            # Despite the swagger schema showing every field optional, this
            # endpoint validates as if the whole resource were being
            # resubmitted -- userIds/roleIds/clientIds/destination/siteIds
            # all come back "required" on a partial update in practice. Echo
            # back the current values unchanged for all of them (access, if
            # it also needs fixing, is applied separately below via the
            # dedicated /users endpoint; roles/clients/destination/siteIds
            # are never modified by this script, see CLAUDE.md).
            update_fields["userIds"] = [u.get("userId") for u in users if u.get("userId")]
            update_fields["roleIds"] = [r["roleId"] for r in roles]
            update_fields["clientIds"] = [c["clientId"] for c in clients]
            update_fields["destination"] = res["destination"]
            update_fields["siteIds"] = res["siteIds"]
            # As of 2026-07-31: also echo back tcp/udp/icmp unchanged.
            # Confirmed live Pangolin behavior -- an update that omits these
            # doesn't just leave them at their swagger-documented "optional"
            # default, it resets udpPortRangeString from blocked back to
            # "all" on the resource being edited. This script never intends
            # to touch ports/icmp, so always re-assert the resource's own
            # current values rather than let the API silently open UDP.
            update_fields["tcpPortRangeString"] = res.get("tcpPortRangeString") or ""
            update_fields["udpPortRangeString"] = res.get("udpPortRangeString") or ""
            update_fields["disableIcmp"] = res.get("disableIcmp", True)
            update_resource(cfg, resource_id, update_fields)
        if needs_access_fix:
            set_resource_users(cfg, resource_id, [target_user_id])
        row["status"] = "OK"
        row["reason"] = reason
    except requests.HTTPError as e:
        row["status"] = "FAIL"
        row["reason"] = f"{e.response.status_code}: {e.response.text[:400]}"
    return row


def write_report(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Normalize Report"

    headers = ["SiteResourceId", "NiceId", "NewNiceId", "Mode", "Destination", "City",
               "TargetEmail", "OldName", "NewName", "Action", "RemovedUsers", "CurrentUsers",
               "Roles", "Clients", "Status", "Reason", "Timestamp"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_colors = {"OK": "C6EFCE", "FAIL": "FFC7CE", "DRY-RUN": "FFEB9C", "SKIPPED": "D9D9D9"}
    for r, row in enumerate(rows, start=2):
        values = [
            row["resource_id"], row["nice_id"], row["new_nice_id"], row["mode"],
            row["destination"], row["city"], row["target_email"], row["old_name"],
            row["new_name"], row["action"], row["removed_users"], row["current_users"],
            row["roles"], row["clients"], row["status"], row["reason"], row["timestamp"],
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=v)
        fill_color = status_colors.get(row["status"])
        if fill_color:
            ws.cell(row=r, column=15).fill = PatternFill("solid", fgColor=fill_color)

    widths = [8, 26, 26, 6, 16, 12, 24, 24, 24, 18, 26, 30, 12, 8, 10, 50, 18]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="config.yml")
    parser.add_argument("--apply", action="store_true",
                         help="actually rename resources / prune access (default: dry run)")
    parser.add_argument("--yes", action="store_true",
                         help="skip the confirmation prompt for a mass --apply")
    parser.add_argument("--resource-id", type=int, action="append", default=None,
                         help="limit to specific siteResourceId(s); repeatable")
    parser.add_argument("--out", default=None, help="report xlsx path (default: auto-named)")
    args = parser.parse_args()

    cfg = load_config(args.config)
    verify_org(cfg)
    sites = get_sites(cfg)

    resources = get_site_resources(cfg)
    if args.resource_id:
        wanted = set(args.resource_id)
        resources = [r for r in resources if r["siteResourceId"] in wanted]
        missing = wanted - {r["siteResourceId"] for r in resources}
        if missing:
            print(f"WARNING: siteResourceId(s) not found: {sorted(missing)}")
    print(f"Found {len(resources)} site resource(s) to evaluate.")

    org_email_index = build_org_email_index(cfg)

    if args.apply and not args.resource_id and not args.yes:
        answer = input(f"About to evaluate {len(resources)} resource(s) and APPLY renames/access "
                        f"changes live. Type 'yes' to continue: ")
        if answer.strip().lower() != "yes":
            sys.exit("Aborted.")

    rows = []
    for res in resources:
        row = process_resource(cfg, sites, res, org_email_index, args.apply)
        rows.append(row)
        if row["action"] != "none" or row["status"] == "SKIPPED":
            print(f"  [{row['status']}] siteResourceId={row['resource_id']} "
                  f"'{row['old_name']}' -> '{row['new_name']}' "
                  f"niceId '{row['nice_id']}' -> '{row['new_nice_id']}' "
                  f"action={row['action']} reason={row['reason'] or '-'}")

    counts = {}
    for row in rows:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
    print(f"\nSummary: {counts}")

    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = args.out or f"private_resources_normalize_{ts}.xlsx"
    write_report(rows, out_path)
    print(f"Report written to: {out_path}")

    if not args.apply:
        print("\nThis was a DRY RUN -- no changes were made. Re-run with --apply to apply them.")


if __name__ == "__main__":
    main()
