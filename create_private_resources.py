#!/usr/bin/env python3
"""
Create Pangolin private (site) resources from a filled-in xlsx request sheet.
Each User Emails entry on a row becomes its OWN site resource (so a row with 3
emails creates 3 resources), each spanning every site in the org (via the
API's siteIds array) for HA. Writes a results report alongside the input file.

Config (config.yml, key-value, nested under 'pangolin'):
    pangolin:
      base_url: "https://pangolin.uop.gr"
      org_slug: "university-of-the-peloponnese"
      api_key: "..."

Requests sheet columns (see pangolin_private_resources_template.xlsx):
    Name | Destination (IP or CIDR) | Ports | Alias | User Emails | Notes

Name: the city (filled in by the Digital Governance Unit, not the requester),
      used as the first segment of the computed resource name.

Alias: optional FQDN (e.g. "app.internal") to reach the resource by name instead
       of IP. Not applicable when Destination is a CIDR range; left blank most
       of the time.

Ports: comma-separated TCP port numbers (e.g. "22,3389"), user-entered.
       UDP and ICMP are always blocked, not user-configurable. Missing/blank
       or any non-numeric token fails that row locally (no API call) rather
       than guessing a port policy.

User Emails: comma-separated. Each email becomes a separate site resource,
             named "<city>-<username>-<vlan>-<tail>" where username is the
             part of the email before "@" with any "." removed (e.g.
             m.katsis -> mkatsis), vlan/tail come from Destination's 10.x.y.z
             octets (host: vlan = x+y zero-padded, tail = unpadded z; CIDR:
             vlan built from whichever octets the prefix fixes, tail "all"
             for a /24 or the literal prefix length otherwise -- identical
             rule to normalize_private_resources.py's, see its module
             docstring). E.g. ktsouvalis@uop.gr at 10.23.2.50 with city
             "patra" -> patra-ktsouvalis-2302-50.
             As of 2026-08-04 (once Pangolin exposed a per-resource `enabled`
             switch), niceId is set deterministically at create time too --
             "<username>-<vlan>-<tail>-<ports>", same formula
             normalize_private_resources.py uses -- regardless of whether the
             email matches a current org user, since the username segment
             only ever depended on the email string, never on a resolved
             account. Before this, niceId was deliberately left to Pangolin's
             own random generation specifically so a still-random niceId
             could signal "nobody's assigned yet" -- with `enabled` that
             trick is no longer needed; disabled state is now the explicit
             signal instead (see the `enabled` bullet below).
             An email that doesn't match any current Pangolin user still gets
             its resource created (name/niceId computed the same way, since
             neither ever depended on the user lookup) but with no user
             attached (userIds: []) and `enabled: false`; reported as
             OK_NO_USER / DRY-RUN_NO_USER rather than silently dropped or
             failed -- its baseline Admin-role default access (see CLAUDE.md)
             shouldn't be reachable by anyone until it has a real owner. Run
             normalize_private_resources.py later, once the account exists,
             to backfill access and flip `enabled` on -- it looks for exactly
             this pattern (correctly-named resource, 0 users) and reverses
             the naming formula to find the matching org user.

Usage:
    python3 create_private_resources.py requests.xlsx --dry-run
    python3 create_private_resources.py requests.xlsx
    python3 create_private_resources.py requests.xlsx --config /path/to/config.yml
"""

import argparse
import datetime
import sys

import requests
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

session = requests.Session()


def load_config(path):
    with open(path, "r") as f:
        cfg = yaml.safe_load(f)
    pg = cfg.get("pangolin", {})
    for key in ("base_url", "org_slug", "api_key"):
        if not pg.get(key):
            sys.exit(f"ERROR: missing 'pangolin.{key}' in {path}")
    pg["base_url"] = pg["base_url"].rstrip("/")
    return pg


def api_headers(cfg):
    return {"Authorization": f"Bearer {cfg['api_key']}", "Content-Type": "application/json"}


def verify_org(cfg):
    resp = session.get(f"{cfg['base_url']}/v1/org/{cfg['org_slug']}", headers=api_headers(cfg))
    resp.raise_for_status()
    org = resp.json()["data"]
    print(f"Org OK: {org.get('name', cfg['org_slug'])} (orgId={cfg['org_slug']})")
    return org


def get_all_pages(cfg, path, data_key, page_size=1000):
    items, page = [], 1
    while True:
        resp = session.get(f"{cfg['base_url']}/v1{path}", headers=api_headers(cfg),
                            params={"pageSize": page_size, "page": page})
        resp.raise_for_status()
        body = resp.json()
        page_items = body["data"][data_key]
        items.extend(page_items)
        total = body["data"].get("pagination", {}).get("total", len(items))
        page += 1
        if len(items) >= total or not page_items:
            break
    return items


def get_sites(cfg):
    sites = get_all_pages(cfg, f"/org/{cfg['org_slug']}/sites", "sites")
    if not sites:
        sys.exit("ERROR: no sites found in this org.")
    return sites


def build_user_index(cfg):
    users = get_all_pages(cfg, f"/org/{cfg['org_slug']}/users", "users")
    index = {}
    for u in users:
        email = (u.get("email") or u.get("user", {}).get("email") or "").lower()
        uid = u.get("id") or u.get("user", {}).get("id")
        if email and uid:
            index[email] = uid
    return index


def parse_tcp_ports(raw_value):
    """Parse the "Ports" column into a normalized comma-separated TCP port
    string, or None if missing/invalid. UDP and ICMP are always blocked, so
    they're not user-configurable inputs (see CLAUDE.md)."""
    raw = str(raw_value).strip() if raw_value else ""
    tokens = [p.strip() for p in raw.split(",") if p.strip()]
    if not tokens or not all(p.isdigit() and 1 <= int(p) <= 65535 for p in tokens):
        return None
    return ",".join(tokens)


def compute_vlan_z(destination):
    """10.x.y.z host -> (vlan, tail), e.g. 10.23.2.50 -> ("2302", "50").
    vlan = x concatenated with y zero-padded to 2 digits; tail is the 4th
    octet, unpadded. (None, None) if not 4 numeric octets or not a 10.x.y.z
    address. Kept identical to normalize_private_resources.py's function of
    the same name -- both scripts must agree on this exactly, since
    normalize_ is what later audits/fixes whatever create_ produces.
    """
    parts = destination.strip().split(".")
    if len(parts) != 4 or not all(p.isdigit() for p in parts) or parts[0] != "10":
        return None, None
    x, y, z = parts[1], parts[2], parts[3]
    return f"{int(x)}{int(y):02d}", str(int(z))


def compute_cidr_vlan_tail(destination):
    """10.x.y.z/prefix -> (vlan, tail), tail "all" for the common /24 case or
    the literal prefix length otherwise (e.g. "/16" -> tail "16"); (None,
    None) if the mask is narrower than /16 (too broad to name a single
    owner's slice). Kept identical to normalize_private_resources.py's
    function of the same name -- see compute_vlan_z's docstring above.
    """
    ip_part, sep, prefix_str = destination.strip().partition("/")
    if not sep or not prefix_str.isdigit():
        return None, None
    prefix = int(prefix_str)
    parts = ip_part.split(".")
    if len(parts) != 4 or not all(p.isdigit() for p in parts) or parts[0] != "10":
        return None, None
    x, y, z = parts[1], parts[2], parts[3]
    fixed_octets = max(0, min(3, prefix // 8 - 1))
    if fixed_octets == 0:
        return None, None
    vlan_tokens = [str(int(x))]
    if fixed_octets >= 2:
        vlan_tokens.append(f"{int(y):02d}")
    if fixed_octets >= 3:
        vlan_tokens.append(str(int(z)))
    tail = "all" if prefix == 24 else str(prefix)
    return "".join(vlan_tokens), tail


def compute_naming_segments(destination, mode):
    """Dispatch to the host or cidr vlan/tail computation for `mode`."""
    if mode == "host":
        return compute_vlan_z(destination)
    if mode == "cidr":
        return compute_cidr_vlan_tail(destination)
    return None, None


def compute_expected_nice_id(username, vlan, tail, ports):
    """"<username>-<vlan>-<tail>-<ports>", ports as sorted "pNNN" tokens.
    Identical formula to normalize_private_resources.py's function of the
    same name -- kept in sync deliberately, see module docstring."""
    return "-".join([username, vlan, tail] + [f"p{p}" for p in sorted(ports, key=int)])


def parse_row(row_num, row, user_index):
    """Return (resolved_reqs, fail_results) for a row.

    Each User Emails entry becomes its own resource request; anything that
    can't produce one (missing emails, bad Ports, unparseable destination,
    unresolved email) becomes a FAIL result instead, without calling the API.
    """
    city, destination, ports_value, alias, emails, notes = row
    if not destination:
        return [], []

    destination = str(destination).strip()
    mode = "cidr" if "/" in destination else "host"
    alias = str(alias).strip() if alias else None
    ports_display = str(ports_value).strip() if ports_value else ""
    tcp_ports = parse_tcp_ports(ports_value)
    city = str(city).strip() if city else ""
    notes_val = str(notes).strip() if notes else None
    vlan, tail = compute_naming_segments(destination, mode)

    raw_emails = [e.strip() for e in str(emails or "").split(",") if e.strip()]

    def make_fail(email, error):
        return {
            "row_num": row_num, "city": city, "name": None, "destination": destination,
            "alias": alias, "tcp_ports": tcp_ports, "email": email, "notes": notes_val,
            "sites": None, "status": "FAIL", "nice_id": None, "enabled": None,
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error": error,
        }

    if not raw_emails:
        return [], [make_fail(None, "no User Emails provided")]

    resolved_reqs, fail_results = [], []
    for raw_email in raw_emails:
        email = raw_email.lower()
        if tcp_ports is None:
            fail_results.append(make_fail(raw_email,
                f"invalid/missing Ports {ports_display!r} "
                f"(expected comma-separated TCP port numbers)"))
            continue
        if vlan is None:
            fail_results.append(make_fail(raw_email,
                f"destination {destination!r} is not a valid 10.x.y.z IPv4 host/CIDR "
                f"(or the CIDR mask is too broad to name a single owner's slice)"))
            continue

        username = email.split("@", 1)[0].replace(".", "")
        resource_name = "-".join([city.lower(), username, vlan, tail])
        nice_id = compute_expected_nice_id(username, vlan, tail, tcp_ports.split(","))

        user_id = user_index.get(email)
        req = {
            "_row_num": row_num,
            "_city": city,
            "_notes": notes_val,
            "_email": raw_email,
            "_user_resolved": user_id is not None,
            "name": resource_name,
            "mode": mode,
            "destination": destination,
            "tcpPortRangeString": tcp_ports,
            "udpPortRangeString": "",
            "disableIcmp": True,
            "roleIds": [],
            "clientIds": [],
            "userIds": [user_id] if user_id is not None else [],
            "niceId": nice_id,
            # Enabled only when a live org account was actually matched --
            # a resource created for an email that doesn't exist in Pangolin
            # yet stays disabled (its baseline Admin-role default access
            # shouldn't be reachable by anyone until it has a real owner);
            # normalize_private_resources.py flips it on once the account
            # exists and access is granted. See module docstring.
            "enabled": user_id is not None,
        }
        if alias:
            req["alias"] = alias
        resolved_reqs.append(req)

    return resolved_reqs, fail_results


def create_site_resource(cfg, sites, req, dry_run):
    payload = {k: v for k, v in req.items() if not k.startswith("_")}
    payload["siteIds"] = [s["siteId"] for s in sites]
    site_names = ", ".join(s["name"] for s in sites)
    no_user = not req["_user_resolved"]

    result = {
        "row_num": req["_row_num"],
        "city": req["_city"],
        "name": req["name"],
        "destination": req["destination"],
        "alias": req.get("alias"),
        "tcp_ports": req["tcpPortRangeString"] or None,
        "email": req["_email"],
        "notes": req["_notes"],
        "sites": site_names,
        "status": None,
        # Set deterministically at request-build time (see module docstring)
        # -- shown here already for a dry run; overwritten with the API's own
        # echoed value below once a live create actually succeeds (should
        # always match what was sent).
        "nice_id": req["niceId"],
        "enabled": req["enabled"],
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "error": (f"no user match: {req['_email']!r} not found among org users -- "
                  f"resource created without access, backfill later via "
                  f"normalize_private_resources.py") if no_user else None,
    }

    if dry_run:
        printable = {k: v for k, v in payload.items() if k not in ("roleIds", "clientIds")}
        print(f"  [DRY-RUN] {printable}")
        result["status"] = "DRY-RUN_NO_USER" if no_user else "DRY-RUN"
        return result

    resp = session.put(f"{cfg['base_url']}/v1/org/{cfg['org_slug']}/site-resource",
                        headers=api_headers(cfg), json=payload)

    if resp.status_code >= 400:
        print(f"  [FAIL] {resp.status_code}: {resp.text}")
        result["status"] = "FAIL"
        result["error"] = f"{resp.status_code}: {resp.text[:500]}"
        return result

    body = resp.json()["data"]
    result["status"] = "OK_NO_USER" if no_user else "OK"
    result["nice_id"] = body["niceId"]
    print(f"  [{'OK-NO-USER' if no_user else 'OK'}] siteResourceId={body['siteResourceId']} "
          f"niceId={body['niceId']} sites=[{site_names}]")
    return result


def write_report(input_path, results):
    wb = load_workbook(input_path)
    if "Results" in wb.sheetnames:
        del wb["Results"]
    ws = wb.create_sheet("Results")

    headers = ["Row", "City", "Name", "Destination", "Alias", "TCP Ports", "Email", "Notes", "Sites",
               "Status", "Nice ID", "Enabled", "Timestamp", "Error"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_colors = {
        "OK": "C6EFCE",
        "FAIL": "FFC7CE",
        "DRY-RUN": "FFEB9C",
        "OK_NO_USER": "FCE4D6",
        "DRY-RUN_NO_USER": "FFF2CC",
    }

    for r, res in enumerate(results, start=2):
        values = [
            res["row_num"], res["city"], res["name"], res["destination"], res["alias"],
            res["tcp_ports"], res["email"], res["notes"], res["sites"],
            res["status"], res["nice_id"], res["enabled"], res["timestamp"], res["error"],
        ]
        for c, v in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=v)
        fill_color = status_colors.get(res["status"])
        if fill_color:
            ws.cell(row=r, column=10).fill = PatternFill("solid", fgColor=fill_color)

    widths = [6, 14, 24, 22, 22, 14, 26, 30, 30, 10, 26, 10, 18, 40]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    out_path = input_path.rsplit(".", 1)[0] + f"_results_{ts}.xlsx"
    wb.save(out_path)
    return out_path


# TODO: add an --update mode to modify an existing site resource identified by
# its niceId from the Results sheet (resolve niceId -> siteResourceId via
# GET /org/{orgId}/site-resources, then POST /site-resource/{siteResourceId}).
# Only creation is supported today.
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", help="path to the filled-in requests xlsx")
    parser.add_argument("--sheet", default="Requests")
    parser.add_argument("--config", default="config.yml")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    cfg = load_config(args.config)
    verify_org(cfg)

    sites = get_sites(cfg)
    print(f"Found {len(sites)} site(s):")
    for s in sites:
        print(f"  - {s['name']} (siteId={s['siteId']}, online={s.get('online')})")

    user_index = build_user_index(cfg)

    wb = load_workbook(args.xlsx_path, data_only=True)
    ws = wb[args.sheet]
    raw_rows = list(ws.iter_rows(min_row=2, max_col=6, values_only=True))

    requests_parsed, upfront_fails = [], []
    for i, row in enumerate(raw_rows, start=2):
        resolved, fails = parse_row(i, row, user_index)
        requests_parsed.extend(resolved)
        upfront_fails.extend(fails)

    print(f"\nParsed {len(raw_rows)} row(s) from '{args.xlsx_path}': "
          f"{len(requests_parsed)} resource(s) to create "
          f"(one per resolved User Emails entry), {len(upfront_fails)} skipped\n")

    all_results = []
    for req in requests_parsed:
        print(f"- row {req['_row_num']}: {req['name']} (niceId={req['niceId']}) "
              f"({req['mode']}: {req['destination']}) "
              f"email={req['_email']} "
              f"alias={req.get('alias') or '-'} "
              f"tcp={req['tcpPortRangeString'] or 'blocked'} "
              f"udp=blocked icmp=blocked "
              f"enabled={req['enabled']}")
        all_results.append(create_site_resource(cfg, sites, req, args.dry_run))

    for fail in upfront_fails:
        print(f"  [FAIL] row {fail['row_num']}: {fail['error']}")
    all_results.extend(upfront_fails)
    all_results.sort(key=lambda r: r["row_num"])

    out_path = write_report(args.xlsx_path, all_results)
    print(f"\nReport written to: {out_path}")

    fails = [r for r in all_results if r["status"] == "FAIL"]
    if fails:
        print(f"WARNING: {len(fails)} row(s)/email(s) failed or were skipped — see report.")

    no_user = [r for r in all_results if r["status"] in ("OK_NO_USER", "DRY-RUN_NO_USER")]
    if no_user:
        print(f"NOTE: {len(no_user)} resource(s) created without a matching user — "
              f"see report; run normalize_private_resources.py later once those "
              f"accounts exist.")


if __name__ == "__main__":
    main()